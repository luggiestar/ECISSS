import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password

from django.http import HttpResponse
from django.shortcuts import redirect, get_object_or_404

from xlsxwriter import Workbook
from ..models import User, Role, Staff


@login_required(login_url='/')
def staff_entry_template(request, templ_name):
    # content-type of response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    template_name = f"{templ_name}"
    name = f"{template_name}.xlsx"
    response['Content-Disposition'] = 'attachment; filename=' + name

    # creating workbook
    wb = Workbook(response, {'in_memory': True})

    # adding sheet
    ws = wb.add_worksheet("sheet1")
    ws.set_column('A:A', 15)
    ws.set_column('B:B', 15)
    ws.set_column('C:C', 20)
    ws.set_column('D:D', 10)
    ws.set_column('E:E', 15)

    # Sheet header, first row
    row_num = 0
    # Add a bold format to use to highlight cells.
    bold = wb.add_format({'bold': 1, 'font_color': 'blue', 'font_name': 'Cambria'})
    bold2 = wb.add_format({'bold': 2, 'font_color': 'red', 'font_name': 'Cambria'})

    # column header names, you can use your own headers here
    columns = ['First Name', 'Last Name', 'email', 'sex', 'phone']

    # write column headers in sheet
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], bold)

    wb.close()
    return response


@login_required(login_url='/')
def upload_staff_entry(request):
    if "GET" == request.method:
        return redirect('staff_list')
    else:
        temp_type = request.POST['type']
        excel_file = request.FILES['staff_entry']
        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb.worksheets[0]

        for rowno, rowval in enumerate(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row - 1), start=2):

            get_username = f"{worksheet.cell(row=rowno, column=1).value}.{worksheet.cell(row=rowno, column=2).value}"
            get_phone = f"0{worksheet.cell(row=rowno, column=5).value}"
            get_role = Role.objects.filter(name='teacher').first()

            try:
                get_staff = get_object_or_404(Staff, user=request.user)
                save_user = User(
                    username=get_username,
                    first_name=worksheet.cell(row=rowno, column=1).value,
                    last_name=worksheet.cell(row=rowno, column=2).value,
                    sex=worksheet.cell(row=rowno, column=4).value,
                    email=worksheet.cell(row=rowno, column=3).value,
                    phone=get_phone,
                    password=make_password(worksheet.cell(row=rowno, column=2).value),
                )

                if save_user.username is None:
                    save_user.delete()
                save_user.save()

                if temp_type == "staff_template":
                    save_staff = Staff(
                        user=save_user,
                        school=get_staff.school,
                        role=get_role
                    )
                    save_staff.save()
                    messages.success(request, f"Staff template uploaded successfully")
                else:
                    messages.success(request, f"User template uploaded successfully")
            except:
                messages.error(request, f"Sometimes went wrong")

        return redirect('staff_list')
